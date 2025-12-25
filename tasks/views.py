from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db import models
from django.http import HttpResponseForbidden, HttpResponse
from .models import Task

import openpyxl
from openpyxl.utils import get_column_letter


@login_required
def my_tasks(request):
    if request.user.is_staff:
        tasks = Task.objects.all()
    else:
        tasks = Task.objects.filter(assigned_to=request.user)

    context = {
        'tasks': tasks,
        'overdue_count': tasks.filter(
            deadline__lt=models.functions.Now()
        ).exclude(status='DONE').count(),
        'due_today_count': tasks.filter(
            deadline=models.functions.Now()
        ).exclude(status='DONE').count(),
        'due_soon_count': tasks.filter(
            deadline__gt=models.functions.Now(),
            deadline__lte=models.functions.Now() + models.DurationField()
        ).exclude(status='DONE').count(),
    }

    return render(request, 'my_tasks.html', context)


@login_required
def update_task(request, task_id):
    task = get_object_or_404(Task, id=task_id)

    if not request.user.is_staff and task.assigned_to != request.user:
        return HttpResponseForbidden("Зөвшөөрөлгүй")

    if request.method == 'POST':
        task.status = request.POST.get('status')
        task.progress = request.POST.get('progress')
        task.deadline = request.POST.get('deadline') or None

        if 'attachment' in request.FILES:
            task.attachment = request.FILES['attachment']

        task.save()
        return redirect('my_tasks')

    return render(request, 'update_task.html', {'task': task})


@login_required
def manager_dashboard(request):
    if not request.user.is_staff:
        return HttpResponseForbidden("Даргын хуудас")

    tasks = Task.objects.all()

    context = {
        'total': tasks.count(),
        'done': tasks.filter(status='DONE').count(),
        'doing': tasks.filter(status='DOING').count(),
        'todo': tasks.filter(status='TODO').count(),
        'overdue': tasks.filter(deadline__lt=models.functions.Now()).exclude(status='DONE').count(),
        'today': tasks.filter(deadline=models.functions.Now()).exclude(status='DONE').count(),
        'soon': tasks.filter(
            deadline__gt=models.functions.Now()
        ).exclude(status='DONE').count(),
        'users': (
            Task.objects
            .values('assigned_to__username')
            .annotate(
                total=models.Count('id'),
                done=models.Count('id', filter=models.Q(status='DONE'))
            )
        )
    }

    return render(request, 'dashboard.html', context)


@login_required
def export_tasks_excel(request):
    if not request.user.is_staff:
        return HttpResponseForbidden("Даргын Excel")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"

    headers = ["Гарчиг", "Ажилтан", "Статус", "Гүйцэтгэл", "Deadline", "Анхааруулга"]
    ws.append(headers)

    for t in Task.objects.select_related('assigned_to').all():
        if t.is_overdue():
            note = "Хоцорсон"
        elif t.is_due_today():
            note = "Өнөөдөр"
        elif t.is_due_soon():
            note = "Ойртож байна"
        else:
            note = ""

        ws.append([
            t.title,
            t.assigned_to.username,
            t.get_status_display(),
            t.progress,
            t.deadline.strftime('%Y-%m-%d') if t.deadline else '',
            note
        ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=tasks_with_notifications.xlsx'
    wb.save(response)

    return response
